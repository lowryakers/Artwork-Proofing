import os
import io
import json
import re
import time
import urllib.request
import openpyxl
from datetime import datetime
from flask import (Flask, render_template, request, redirect,
                   url_for, flash, jsonify, send_file)

import proof_engine

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'prodough-proof-site-2024-local')
app.config['MAX_CONTENT_LENGTH'] = 500 * 1024 * 1024  # 500 MB

BASE_DIR              = os.path.dirname(os.path.abspath(__file__))
UPLOAD_DIR            = os.path.join(BASE_DIR, 'uploads')
GTIN_STORE_PATH       = os.path.join(UPLOAD_DIR, 'gtin_store.json')
GTIN_SHEET_CFG_PATH   = os.path.join(UPLOAD_DIR, 'gtin_sheet_config.json')
os.makedirs(UPLOAD_DIR, exist_ok=True)

_sheet_cache: dict = {'rows': None, 'fetched_at': 0.0, 'url': ''}
_SHEET_CACHE_TTL = 300  # seconds

ALLOWED_EXTS = {'.pdf', '.ai', '.png', '.jpg', '.jpeg', '.tiff', '.tif', '.eps'}

CHECK_LABELS = {
    'gtin':     'GTIN / Barcode',
    'nfp':      'Front Call-Out vs NFP',
    'eyemark':  'Eyemark Contrast',
    'spelling': 'Spelling / Brand Name',
    'fda':      'FDA Audit Risk',
    'wind':     'Wind Direction',
    'specs':    'Print Specs',
}


@app.template_filter('basename_filter')
def basename_filter(path):
    return os.path.basename(path) if path else ''


# ── GTIN store ────────────────────────────────────────────────────────────────

def _load_gtin_store():
    if os.path.exists(GTIN_STORE_PATH):
        try:
            with open(GTIN_STORE_PATH) as f:
                return json.load(f)
        except Exception:
            pass
    return None


def _save_gtin_store(rows: list, source_filename: str) -> dict:
    for row in rows:
        gtin = str(row.get('gtin', ''))
        if not gtin:
            row['_valid'] = False
            row['_error'] = 'Missing GTIN'
        elif not gtin.isdigit():
            row['_valid'] = False
            row['_error'] = 'Non-numeric characters'
        elif len(gtin) != 12:
            row['_valid'] = False
            row['_error'] = f'{len(gtin)} digits (expect 12)'
        else:
            row['_valid'] = True
            row['_error'] = None
    store = {
        'uploaded_at': datetime.now().isoformat(),
        'filename': source_filename,
        'row_count': len(rows),
        'error_count': sum(1 for r in rows if not r.get('_valid')),
        'rows': rows,
    }
    with open(GTIN_STORE_PATH, 'w') as f:
        json.dump(store, f)
    return store


# ── Google Sheets sync ───────────────────────────────────────────────────────

_DEFAULT_CFG_PATH = os.path.join(BASE_DIR, 'gtin_default_config.json')

def _load_sheet_config() -> dict:
    # 1. Runtime config file (set via UI, wiped on redeploy)
    if os.path.exists(GTIN_SHEET_CFG_PATH):
        try:
            with open(GTIN_SHEET_CFG_PATH) as f:
                cfg = json.load(f)
                if cfg.get('sheet_url'):
                    return cfg
        except Exception:
            pass
    # 2. Environment variable
    env_url = os.environ.get('GTIN_SHEET_URL', '').strip()
    if env_url:
        return {'sheet_url': env_url}
    # 3. Default config baked into the image (survives redeploys)
    if os.path.exists(_DEFAULT_CFG_PATH):
        try:
            with open(_DEFAULT_CFG_PATH) as f:
                cfg = json.load(f)
                if cfg.get('sheet_url'):
                    return cfg
        except Exception:
            pass
    return {}


def _save_sheet_config(cfg: dict):
    with open(GTIN_SHEET_CFG_PATH, 'w') as f:
        json.dump(cfg, f)


def _sheet_url_to_csv(url: str) -> str:
    """Convert any Google Sheets share/edit/published URL to a direct CSV export URL."""
    import re
    if 'export?format=csv' in url or 'output=csv' in url:
        return url
    # Published-to-web format: /spreadsheets/d/e/LONG_KEY/pubhtml
    m = re.search(r'(https://docs\.google\.com/spreadsheets/d/e/[a-zA-Z0-9_-]+)/pub', url)
    if m:
        return f'{m.group(1)}/pub?output=csv'
    # Standard share/edit format: /spreadsheets/d/SHEET_ID/...
    m = re.search(r'/spreadsheets/d/([a-zA-Z0-9_-]+)', url)
    if m and m.group(1) != 'e':
        sheet_id = m.group(1)
        gid_m = re.search(r'[#&?]gid=(\d+)', url)
        gid_part = f'&gid={gid_m.group(1)}' if gid_m else ''
        return f'https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=csv{gid_part}'
    raise ValueError('Not a recognized Google Sheets URL. Paste the share/edit URL from your browser address bar.')


def _fetch_sheet_rows(csv_url: str) -> list:
    """Download and parse the Master SKU Google Sheet CSV.
    Returns unified rows used for both GTIN checking and spec validation.
    """
    import csv as _csv
    req = urllib.request.Request(csv_url, headers={'User-Agent': 'Mozilla/5.0'})
    with urllib.request.urlopen(req, timeout=15) as resp:
        content = resp.read().decode('utf-8-sig')
    reader = _csv.DictReader(io.StringIO(content))
    rows = []
    for row in reader:
        d = {k.strip().lower(): (v.strip() if v else '') for k, v in row.items() if k}

        def _get(*keys):
            for k in keys:
                v = d.get(k, '')
                if v and v.lower() not in ('nan', 'none'):
                    return v
            return ''

        flavor   = _get('flavor', 'product name', 'name', 'updated naming', 'product title')
        sku      = _get('sku', 'variant sku')
        gtin_raw = _get('gtin/barcode', 'gtin/barcode#', 'gtin / upc', 'gtin/upc',
                        'gtin', 'upc', 'barcode')
        pkg      = _get('packaging type', 'package', 'type')
        mat      = _get('material', 'material order', 'substrate')
        zipper   = _get('zipper')
        print_p  = _get('print', 'print process')

        def _to_float(v):
            if not v:
                return None
            # Strip common unit suffixes before converting (e.g. "254mm" → 254.0)
            v_clean = re.sub(r'\s*(?:mm|cm|in|inch(?:es)?|ft|g|oz|lbs?)\s*$', '',
                             str(v).strip(), flags=re.IGNORECASE)
            try:
                return float(v_clean) if v_clean else None
            except (ValueError, TypeError):
                return None

        trim_length = _to_float(_get('trim length', 'length mm', 'trim l', 'height mm', 'trim h'))
        trim_width  = _to_float(_get('trim width',  'width mm',  'trim w', 'w mm'))
        gusset      = _to_float(_get('gusset dimension', 'gusset'))
        front_panel = _to_float(_get('front panel dimension', 'front panel'))

        wind_raw    = _get('wind direction', 'wind', 'winding')
        wind        = re.sub(r'[^\d]', '', wind_raw)[:1]

        pms_colors  = _get('pms spot colors', 'pms colors', 'spot colors', 'pantone')
        hex_colors  = _get('hex spot colors', 'hex colors')
        eye_mark    = _get('eye mark color', 'eye mark', 'eyemark color', 'eyemark')

        die_raw      = _get('die line required', 'die line', 'die lines').strip().lower()
        die_required = die_raw in ('yes', 'y', 'true', '1')

        # Normalize GTIN — Excel stores long integers as floats (e.g. 850012345678.0)
        gtin_str = str(gtin_raw).strip().replace(' ', '')
        if '.' in gtin_str and gtin_str.replace('.', '').isdigit():
            try:
                gtin_str = str(int(float(gtin_str)))
            except (ValueError, OverflowError):
                pass

        if not flavor and not gtin_str:
            continue  # skip blank rows

        gtin_valid = bool(gtin_str) and gtin_str.isdigit() and len(gtin_str) == 12

        rows.append({
            'flavor':            str(flavor).strip().lower(),
            'sku':               str(sku).strip(),
            'gtin':              gtin_str,
            '_valid':            gtin_valid,
            '_error':            None if gtin_valid else (
                'Missing GTIN'       if not gtin_str else
                f'{len(gtin_str)} digits (expect 12)' if gtin_str.isdigit() else
                'Non-numeric'
            ),
            'packaging_type':    pkg,
            'material':          mat,
            'zipper':            zipper,
            'print_process':     print_p,
            'trim_length_mm':    trim_length,
            'trim_height_mm':    trim_length,   # alias — proof_engine uses trim_height_mm
            'trim_width_mm':     trim_width,
            'gusset_mm':         gusset,
            'front_panel_mm':    front_panel,
            'wind_direction':    wind,
            'pms_spot_colors':   pms_colors,
            'hex_spot_colors':   hex_colors,
            'eye_mark_color':    eye_mark,
            'die_line_required': die_required,
        })
    return rows


def _get_sheet_gtin_rows(force: bool = False) -> list:
    """Return GTIN rows from the synced Google Sheet, with a 5-minute in-memory cache."""
    global _sheet_cache
    cfg = _load_sheet_config()
    url = cfg.get('sheet_url', '')
    if not url:
        return []
    now = time.time()
    if (not force and _sheet_cache['url'] == url
            and _sheet_cache['rows'] is not None
            and now - _sheet_cache['fetched_at'] < _SHEET_CACHE_TTL):
        return _sheet_cache['rows']
    try:
        csv_url = _sheet_url_to_csv(url)
        rows = _fetch_sheet_rows(csv_url)
        _sheet_cache = {'rows': rows, 'fetched_at': now, 'url': url}
        cfg['last_synced'] = datetime.now().isoformat()
        cfg['row_count'] = len(rows)
        _save_sheet_config(cfg)
        return rows
    except Exception as exc:
        _sheet_cache['last_error'] = str(exc)
        return _sheet_cache['rows'] if _sheet_cache['rows'] is not None else []




# ── Pages ─────────────────────────────────────────────────────────────────────

@app.route('/')
def landing():
    proof_engine.load_jobs_from_disk()
    return render_template('landing.html', jobs=proof_engine.list_jobs())


@app.route('/proof')
def prodough_proof():
    return render_template('proof.html', jobs=proof_engine.list_jobs())


@app.route('/brand')
def brand_page():
    return render_template('brand.html', jobs=proof_engine.list_jobs())


@app.route('/gtin')
def gtin_page():
    store = _load_gtin_store()
    sheet_cfg = _load_sheet_config()
    # If a sheet is configured, show the live sheet rows in the table
    if sheet_cfg.get('sheet_url') and not store:
        rows = _get_sheet_gtin_rows()
        if rows:
            store = _save_gtin_store(rows, 'Google Sheets (auto-synced)')
        elif _sheet_cache.get('last_error'):
            flash(f'Sheet sync failed: {_sheet_cache["last_error"]}', 'danger')
    env_url = os.environ.get('GTIN_SHEET_URL', '')
    return render_template('gtin.html', store=store, sheet_cfg=sheet_cfg, env_sheet_url=env_url)


@app.route('/specs')
def specs_page():
    return redirect(url_for('gtin_page'))


@app.route('/upload', methods=['POST'])
def upload():
    artwork_files = request.files.getlist('artwork')
    gtin_file     = request.files.get('gtin_list')

    if not artwork_files or all(f.filename == '' for f in artwork_files):
        flash('Please select at least one artwork file.', 'danger')
        return redirect(url_for('prodough_proof'))

    # GTIN data: uploaded file takes priority; fall back to synced Google Sheet
    gtin_rows = []
    if gtin_file and gtin_file.filename:
        try:
            gtin_rows = _parse_gtin_list(gtin_file.read(), gtin_file.filename)
            if gtin_rows:
                _save_gtin_store(gtin_rows, gtin_file.filename)
            else:
                flash('The uploaded GTIN file appears empty — falling back to synced sheet data.', 'warning')
        except Exception as exc:
            flash(f'Could not read the uploaded GTIN file ({exc}) — falling back to synced sheet data.', 'warning')

    if not gtin_rows:
        gtin_rows = _get_sheet_gtin_rows()  # use cache; background-refresh below
        if not gtin_rows and _sheet_cache.get('last_error'):
            flash(f'Google Sheet sync failed: {_sheet_cache["last_error"]}', 'danger')

    # Kick off a background sheet refresh so the *next* proof run has fresh data
    # without blocking this one on a network round-trip.
    if _load_sheet_config().get('sheet_url'):
        import threading as _t
        _t.Thread(target=_get_sheet_gtin_rows, kwargs={'force': True}, daemon=True).start()

    if not gtin_rows:
        flash(
            'No SKU data available. Connect your Master SKU Sheet on the SKU Master page '
            'so the barcode and spec checks have a master list to match against.',
            'warning',
        )

    wind_direction = request.form.get('wind_direction', '').strip()
    proof_type     = request.form.get('proof_type', 'press').strip()
    prodough_config = {
        'brand_mode':    'prodough',
        'wind_direction': wind_direction,
        'proof_type':    proof_type,
    }

    spec_rows = gtin_rows  # unified sheet — same rows serve both checks

    job_id  = proof_engine.create_job([f.filename for f in artwork_files if f.filename],
                                      brand_config=prodough_config)
    job_dir = os.path.join(UPLOAD_DIR, job_id)
    os.makedirs(job_dir, exist_ok=True)

    saved, skipped = [], []
    for f in artwork_files:
        if not f.filename:
            continue
        ext = os.path.splitext(f.filename)[1].lower()
        if ext not in ALLOWED_EXTS:
            skipped.append(f.filename)
            continue
        safe = _safe_name(f.filename)
        dest = os.path.join(job_dir, safe)
        f.save(dest)
        saved.append(dest)

    if skipped:
        flash(f'Skipped unsupported file(s): {", ".join(skipped)}', 'warning')
    if not saved:
        flash('No supported artwork files were uploaded.', 'danger')
        return redirect(url_for('prodough_proof'))

    proof_engine.start_job(job_id, saved, gtin_rows, job_dir, brand_config=prodough_config,
                           spec_rows=spec_rows)
    return redirect(url_for('result', job_id=job_id))


@app.route('/brand/upload', methods=['POST'])
def brand_upload():
    brand_name    = request.form.get('brand_name', '').strip()
    packaging_type = request.form.get('packaging_type', 'other')
    artwork_files = request.files.getlist('artwork')
    gtin_file     = request.files.get('gtin_list')

    if not artwork_files or all(f.filename == '' for f in artwork_files):
        flash('Please select at least one artwork file.', 'danger')
        return redirect(url_for('brand_page'))

    if not gtin_file or not gtin_file.filename:
        flash('A GTIN list is required for Brand Proof. Please upload your brand\'s SKU & GTIN spreadsheet.', 'danger')
        return redirect(url_for('brand_page'))

    try:
        gtin_rows = _parse_gtin_list(gtin_file.read(), gtin_file.filename)
    except Exception as exc:
        flash(f'Could not read the GTIN file: {exc}', 'danger')
        return redirect(url_for('brand_page'))

    if not gtin_rows:
        flash('The uploaded GTIN file appears empty. Please check the file and try again.', 'warning')

    wind_direction = request.form.get('wind_direction', '').strip()
    proof_type     = request.form.get('proof_type', 'press').strip()
    brand_config = {
        'brand_mode':      'generic',
        'brand_name':      brand_name,
        'packaging_type':  packaging_type,
        'wind_direction':  wind_direction,
        'proof_type':      proof_type,
    }

    spec_rows = gtin_rows  # unified sheet — same rows serve both checks

    job_id  = proof_engine.create_job([f.filename for f in artwork_files if f.filename],
                                      brand_config=brand_config)
    job_dir = os.path.join(UPLOAD_DIR, job_id)
    os.makedirs(job_dir, exist_ok=True)

    saved, skipped = [], []
    for f in artwork_files:
        if not f.filename:
            continue
        ext = os.path.splitext(f.filename)[1].lower()
        if ext not in ALLOWED_EXTS:
            skipped.append(f.filename)
            continue
        safe = _safe_name(f.filename)
        dest = os.path.join(job_dir, safe)
        f.save(dest)
        saved.append(dest)

    if skipped:
        flash(f'Skipped unsupported file(s): {", ".join(skipped)}', 'warning')
    if not saved:
        flash('No supported artwork files were uploaded.', 'danger')
        return redirect(url_for('brand_page'))

    proof_engine.start_job(job_id, saved, gtin_rows, job_dir, brand_config=brand_config,
                           spec_rows=spec_rows)
    return redirect(url_for('result', job_id=job_id))


@app.route('/result/<job_id>')
def result(job_id):
    job = proof_engine.get_job(job_id)
    if not job:
        flash('Job not found.', 'danger')
        return redirect(url_for('landing'))
    # Build verify_map inline — items that say "verify" or "manually" in their message
    verify_map = {}
    for r in job.get('results', []):
        items = []
        for check_name, check in r.get('checks', {}).items():
            for issue in check.get('issues', []):
                msg = issue.get('message', '')
                if 'verify' in msg.lower() or 'manually' in msg.lower():
                    items.append({'check_label': CHECK_LABELS.get(check_name, check_name),
                                  'severity': issue['severity'], 'message': msg})
            for note in check.get('notes', []):
                if 'verify' in note.lower() or 'manually' in note.lower():
                    items.append({'check_label': CHECK_LABELS.get(check_name, check_name),
                                  'severity': 'info', 'message': note})
        verify_map[r['filename']] = items
    return render_template('result.html', job=job, job_id=job_id, verify_map=verify_map)


@app.route('/debug/ocr/<job_id>/<path:filename>')
def debug_ocr(job_id, filename):
    """Show the raw OCR text extracted from a specific file in a job."""
    job = proof_engine.get_job(job_id)
    if not job:
        return 'Job not found', 404
    for r in job.get('results', []):
        if r.get('filename') == filename:
            ocr = r.get('ocr_text', '(not stored)')
            return f'<pre style="white-space:pre-wrap;font-size:13px;padding:1rem">' \
                   f'FILE: {filename}\nJOB: {job_id}\nOCR LENGTH: {len(ocr)} chars\n\n' \
                   f'{"="*60}\n{ocr}\n{"="*60}</pre>'
    return f'File {filename!r} not found in job {job_id}', 404


@app.route('/debug/ocr/<job_id>')
def debug_ocr_list(job_id):
    """List all files in a job with links to their OCR text."""
    job = proof_engine.get_job(job_id)
    if not job:
        return 'Job not found', 404
    links = ''.join(
        f'<li><a href="/debug/ocr/{job_id}/{r["filename"]}">{r["filename"]}</a></li>'
        for r in job.get('results', []) if not r.get('error')
    )
    return f'<ul>{links}</ul>'


@app.route('/viewer/<job_id>')
def viewer(job_id):
    return redirect(url_for('result', job_id=job_id))


@app.route('/summary/<job_id>')
def summary(job_id):
    job = proof_engine.get_job(job_id)
    if not job:
        flash('Job not found.', 'danger')
        return redirect(url_for('landing'))
    if job['status'] != 'done':
        return redirect(url_for('result', job_id=job_id))
    return render_template('summary.html', job=job, job_id=job_id)


@app.route('/history')
def history():
    proof_engine.load_jobs_from_disk()
    jobs = proof_engine.list_jobs()
    return render_template('history.html', jobs=jobs)



# ── API ───────────────────────────────────────────────────────────────────────

@app.route('/api/status/<job_id>')
def api_status(job_id):
    job = proof_engine.get_job(job_id)
    if not job:
        return jsonify({'error': 'not found'}), 404
    return jsonify({
        'status':       job['status'],
        'progress':     job['progress'],
        'current_file': job['current_file'],
        'error':        job['error'],
    })


@app.route('/api/result/<job_id>')
def api_result(job_id):
    job = proof_engine.get_job(job_id)
    if not job:
        return jsonify({'error': 'not found'}), 404
    return jsonify(job)


@app.route('/api/gtin-sheet', methods=['POST'])
def api_gtin_sheet_save():
    data = request.get_json() or {}
    url = data.get('url', '').strip()
    if url:
        try:
            _sheet_url_to_csv(url)  # validate it's a Sheets URL
        except ValueError as e:
            return jsonify({'ok': False, 'error': str(e)}), 400
    cfg = _load_sheet_config()
    cfg['sheet_url'] = url
    cfg.pop('last_synced', None)
    cfg.pop('row_count', None)
    _save_sheet_config(cfg)
    global _sheet_cache
    _sheet_cache = {'rows': None, 'fetched_at': 0.0, 'url': ''}
    return jsonify({'ok': True})


@app.route('/api/gtin-sheet/sync', methods=['POST'])
def api_gtin_sheet_sync():
    cfg = _load_sheet_config()
    if not cfg.get('sheet_url'):
        return jsonify({'ok': False, 'error': 'No sheet URL configured'}), 400
    try:
        rows = _get_sheet_gtin_rows(force=True)
        if rows:
            _save_gtin_store(rows, 'Google Sheets (auto-synced)')
        cfg = _load_sheet_config()
        return jsonify({'ok': True, 'row_count': len(rows), 'last_synced': cfg.get('last_synced', '')})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500



@app.route('/api/dismiss/<job_id>', methods=['POST'])
def api_dismiss(job_id):
    data = request.get_json()
    if not data:
        return jsonify({'error': 'no data'}), 400
    ok = proof_engine.set_dismissal(
        job_id,
        data.get('filename', ''),
        data.get('check', ''),
        int(data.get('index', 0)),
        bool(data.get('dismissed', True)),
    )
    if not ok:
        return jsonify({'error': 'job not found'}), 404
    return jsonify({'ok': True})


@app.route('/report/<job_id>')
def download_report(job_id):
    job = proof_engine.get_job(job_id)
    if not job or job['status'] != 'done':
        flash('Report not ready — job must complete first.', 'danger')
        return redirect(url_for('landing'))
    brand_name = job.get('brand_config', {}).get('brand_name', 'ProDough')
    buf = _generate_report(job, brand_name=brand_name)
    safe_brand = _safe_name(brand_name).replace(' ', '_')
    fname = f'{safe_brand}_Proof_{job_id}.xlsx'
    return send_file(
        buf, download_name=fname, as_attachment=True,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )


@app.route('/image/<job_id>/<path:filename>')
def serve_image(job_id, filename):
    job_dir  = os.path.realpath(os.path.join(UPLOAD_DIR, job_id))
    img_path = os.path.realpath(os.path.join(job_dir, filename))
    if not img_path.startswith(job_dir):
        return 'Forbidden', 403
    if not os.path.exists(img_path):
        return 'Not found', 404
    return send_file(img_path, mimetype='image/png')


# ── Helpers ───────────────────────────────────────────────────────────────────

def _safe_name(name: str) -> str:
    keep = set('abcdefghijklmnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ0123456789._- ')
    return ''.join(c if c in keep else '_' for c in name)


def _parse_gtin_list(data: bytes, filename: str) -> list:
    rows = []
    if filename.lower().endswith('.csv'):
        import csv
        reader = csv.DictReader(io.TextIOWrapper(io.BytesIO(data), errors='replace'))
        for row in reader:
            rows.append({k.strip().lower(): v.strip() for k, v in row.items()})
    else:
        wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
        for ws in wb.worksheets:
            headers = None
            for row in ws.iter_rows(values_only=True):
                if headers is None:
                    if row and any(row):
                        headers = [str(c).strip().lower() if c else '' for c in row]
                    continue
                if not any(row):
                    continue
                d = dict(zip(headers, row))
                gtin_raw = d.get('gtin/barcode#') or d.get('gtin') or d.get('barcode') or d.get('upc') or ''
                flavor   = d.get('flavor') or d.get('product name') or d.get('name') or ''
                sku      = d.get('sku') or ''
                # Excel stores long numbers as floats (e.g. 850012345678.0) — convert via int first
                if isinstance(gtin_raw, (int, float)):
                    gtin_str = str(int(gtin_raw))
                else:
                    gtin_str = str(gtin_raw).strip().replace(' ', '')
                    if '.' in gtin_str and gtin_str.replace('.', '').replace('-', '').isdigit():
                        try:
                            gtin_str = str(int(float(gtin_str)))
                        except (ValueError, OverflowError):
                            pass
                rows.append({
                    'gtin':   gtin_str,
                    'flavor': str(flavor).strip().lower(),
                    'sku':    str(sku).strip(),
                })
    return [r for r in rows if r.get('gtin') and r['gtin'] not in ('', 'nan')]


def _generate_report(job: dict, brand_name: str = 'ProDough') -> io.BytesIO:
    from openpyxl.styles import Font, PatternFill, Alignment

    dismissals = job.get('dismissals', {})

    def is_dismissed(filename, check_name, idx):
        return dismissals.get(f"{filename}|{check_name}|{idx}", False)

    hdr_font  = Font(bold=True, color='FFFFFF')
    hdr_blue  = PatternFill('solid', fgColor='1A56A0')
    hdr_gray  = PatternFill('solid', fgColor='6C757D')
    fill_crit = PatternFill('solid', fgColor='FFD0D0')
    fill_warn = PatternFill('solid', fgColor='FFF3CD')
    fill_ok   = PatternFill('solid', fgColor='D4EDDA')
    wrap      = Alignment(wrap_text=True, vertical='top')
    center    = Alignment(horizontal='center', vertical='top')

    wb = openpyxl.Workbook()

    # ── Sheet 1: Summary ──────────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = 'Summary'
    ws1.merge_cells('A1:E1')
    ws1['A1'] = f'{brand_name} Artwork Proof Report'
    ws1['A1'].font = Font(bold=True, size=14)
    ws1['A2'] = f'Job: {job["id"]}'
    ws1['B2'] = f'Date: {job["created"][:10]}'
    ws1['C2'] = f'Files proofed: {len(job["results"])}'
    ws1.append([])
    for col, hdr in enumerate(['File', 'Overall Status', 'Critical', 'Warnings', 'Notes'], 1):
        c = ws1.cell(row=4, column=col, value=hdr)
        c.font = hdr_font
        c.fill = hdr_blue
        c.alignment = center

    for r in job['results']:
        ws1.append([r['filename'], r.get('severity', 'error').upper(),
                    r.get('critical_count', 0), r.get('warning_count', 0), r.get('info_count', 0)])
        sev = r.get('severity', 'error')
        f = fill_crit if sev == 'critical' else fill_warn if sev == 'warning' else fill_ok if sev == 'clean' else None
        if f:
            for col in range(1, 6):
                ws1.cell(row=ws1.max_row, column=col).fill = f

    ws1.column_dimensions['A'].width = 45
    for col in ['B', 'C', 'D', 'E']:
        ws1.column_dimensions[col].width = 14

    # ── Sheet 2: Issues to Fix (non-dismissed) ────────────────────────────────
    ws2 = wb.create_sheet('Issues to Fix')
    for col, hdr in enumerate(['File', 'Check', 'Severity', 'Issue / Required Action'], 1):
        c = ws2.cell(row=1, column=col, value=hdr)
        c.font = hdr_font
        c.fill = hdr_blue
        c.alignment = center

    row_idx = 2
    for r in job['results']:
        for check_name, check in r.get('checks', {}).items():
            for i, issue in enumerate(check.get('issues', [])):
                if is_dismissed(r['filename'], check_name, i):
                    continue
                sev = issue['severity']
                vals = [r['filename'], CHECK_LABELS.get(check_name, check_name),
                        sev.upper(), issue['message']]
                for col, val in enumerate(vals, 1):
                    c = ws2.cell(row=row_idx, column=col, value=val)
                    c.fill = fill_crit if sev == 'critical' else fill_warn if sev == 'warning' else PatternFill()
                    c.alignment = wrap
                row_idx += 1

    ws2.column_dimensions['A'].width = 40
    ws2.column_dimensions['B'].width = 22
    ws2.column_dimensions['C'].width = 12
    ws2.column_dimensions['D'].width = 80

    # ── Sheet 3: Reviewed / Dismissed ─────────────────────────────────────────
    if dismissals:
        ws3 = wb.create_sheet('Reviewed OK')
        for col, hdr in enumerate(['File', 'Check', 'Severity', 'Issue (Reviewed — No Action Needed)'], 1):
            c = ws3.cell(row=1, column=col, value=hdr)
            c.font = hdr_font
            c.fill = hdr_gray
            c.alignment = center

        row_idx = 2
        for r in job['results']:
            for check_name, check in r.get('checks', {}).items():
                for i, issue in enumerate(check.get('issues', [])):
                    if not is_dismissed(r['filename'], check_name, i):
                        continue
                    vals = [r['filename'], CHECK_LABELS.get(check_name, check_name),
                            issue['severity'].upper(), issue['message']]
                    for col, val in enumerate(vals, 1):
                        ws3.cell(row=row_idx, column=col, value=val).alignment = wrap
                    row_idx += 1

        ws3.column_dimensions['A'].width = 40
        ws3.column_dimensions['B'].width = 22
        ws3.column_dimensions['C'].width = 12
        ws3.column_dimensions['D'].width = 80

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
